import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
import warnings
warnings.filterwarnings('ignore')

print("正在读取数据并生成完整的Excel分析报告...")

# ==================== 配置 ====================
# 文件路径
FILE_PREV = '2025-7-2.xlsx'  # 上次考试
FILE_CURR = '2025-7-3.xlsx'  # 本次考试

# 科目配置 - 支持两种命名方式
SUBJECTS = ['语文', '数学', '英语', '科学', '社会']  # 报告中显示的全称
SUBJECTS_SHORT = ['语', '数', '英', '科', '社']  # 单字简写
SUBJECTS_120 = ['语文', '数学', '英语', '科学']  # 满分120的科目
SUBJECTS_100 = ['社会']  # 满分100的科目

# 评分标准（按科目满分制）
PASS_LINES = {
    '语文': 72, '数学': 72, '英语': 72, '科学': 72, '社会': 60  # 及格线 (60%)
}
EXCEL_LINES = {
    '语文': 108, '数学': 108, '英语': 108, '科学': 108, '社会': 90  # 优秀线 (90%)
}

# 总分满分（4科×120 + 1科×100）
TOTAL_FULL_SCORE = 580

# ==================== 数据读取 ====================
def read_exam_data(filepath):
    """读取考试数据，合并各科目sheet，自动检测命名方式"""
    import openpyxl
    all_data = {}

    # 先获取所有sheet名称来判断命名方式
    wb = openpyxl.load_workbook(filepath, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    # 判断是简写还是全称
    use_short = '总' in sheet_names
    total_sheet = '总' if use_short else '总分'

    # 建立科目映射
    if use_short:
        subject_map = dict(zip(SUBJECTS_SHORT, SUBJECTS))  # 简写->全称
        sheet_list = SUBJECTS_SHORT
    else:
        subject_map = dict(zip(SUBJECTS, SUBJECTS))  # 全称->全称
        sheet_list = SUBJECTS

    # 读取总分sheet获取学生基本信息
    try:
        df_total = pd.read_excel(filepath, sheet_name=total_sheet)
        df_total.columns = ['学号', '姓名', '总分', '总分名次', '系数']
        # 统一学号格式：先转数值再转整数再转字符串
        df_total['学号'] = pd.to_numeric(df_total['学号'], errors='coerce').fillna(0).astype(int).astype(str)
        df_total['总分'] = pd.to_numeric(df_total['总分'], errors='coerce')
        df_total['总分名次'] = pd.to_numeric(df_total['总分名次'], errors='coerce')
        all_data['总分'] = df_total[['学号', '姓名', '总分', '总分名次']].copy()
    except Exception as e:
        print(f"读取总分sheet失败: {e}")
        return None

    # 读取各科目sheet
    for sheet_name, subject in zip(sheet_list, SUBJECTS):
        try:
            df_subj = pd.read_excel(filepath, sheet_name=sheet_name)
            # 列名为：学号、姓名、总分（这里是该科成绩）、名次、系数
            df_subj.columns = ['学号', '姓名', f'{subject}', f'{subject}名次', '系数']
            # 统一学号格式
            df_subj['学号'] = pd.to_numeric(df_subj['学号'], errors='coerce').fillna(0).astype(int).astype(str)
            df_subj[subject] = pd.to_numeric(df_subj[subject], errors='coerce')
            df_subj[f'{subject}名次'] = pd.to_numeric(df_subj[f'{subject}名次'], errors='coerce')
            all_data[subject] = df_subj[['学号', '姓名', subject, f'{subject}名次']].copy()
        except Exception as e:
            print(f"读取{subject}({sheet_name})sheet失败: {e}")

    # 合并所有数据
    merged = all_data['总分'].copy()
    for subject in SUBJECTS:
        if subject in all_data:
            merged = pd.merge(merged, all_data[subject][['学号', subject, f'{subject}名次']],
                            on='学号', how='left')

    return merged

print("读取上次考试数据...")
df_prev = read_exam_data(FILE_PREV)
print(f"  上次考试学生数: {len(df_prev)}")

print("读取本次考试数据...")
df_curr = read_exam_data(FILE_CURR)
print(f"  本次考试学生数: {len(df_curr)}")

# ==================== 数据合并与变化计算 ====================
print("合并数据并计算变化...")

# 合并两次考试数据
merged_df = pd.merge(
    df_curr,
    df_prev,
    on='学号',
    suffixes=('_本次', '_上次')
)

print(f"  匹配学生数: {len(merged_df)}")

# 计算各科变化
for subject in SUBJECTS:
    merged_df[f'{subject}_变化'] = merged_df[f'{subject}_本次'] - merged_df[f'{subject}_上次']
    merged_df[f'{subject}名次_变化'] = merged_df[f'{subject}名次_上次'] - merged_df[f'{subject}名次_本次']

# 计算总分变化
merged_df['总分_变化'] = merged_df['总分_本次'] - merged_df['总分_上次']
merged_df['总分名次_变化'] = merged_df['总分名次_上次'] - merged_df['总分名次_本次']

# ==================== 学生个人分析函数 ====================
def analyze_student(row):
    """分析单个学生的成绩变化"""
    analysis_parts = []
    total_change = row['总分_变化']
    rank_change = row['总分名次_变化']

    # 总分和排名变化
    analysis_parts.append(f"总分{row['总分_上次']:.0f}→{row['总分_本次']:.0f}")
    if pd.notna(rank_change):
        analysis_parts.append(f"名次{int(row['总分名次_上次'])}→{int(row['总分名次_本次'])}")

    # 判断阈值（总分满分580分）
    if total_change > 30:  # 约5%以上的提升
        analysis_parts.append("成绩明显进步")
    elif total_change > 0:
        analysis_parts.append("成绩进步")
    elif total_change > -30:
        analysis_parts.append("成绩略有下滑")
    else:
        analysis_parts.append("成绩明显下滑")

    summary = "，".join(analysis_parts)

    # 波动原因推测
    reasons = []
    subject_changes = {s: row[f'{s}_变化'] for s in SUBJECTS}

    # 找出变化最大和最小的科目
    max_subject = max(subject_changes, key=lambda x: subject_changes[x] if pd.notna(subject_changes[x]) else -999)
    max_change = subject_changes[max_subject] if pd.notna(subject_changes[max_subject]) else 0
    min_subject = min(subject_changes, key=lambda x: subject_changes[x] if pd.notna(subject_changes[x]) else 999)
    min_change = subject_changes[min_subject] if pd.notna(subject_changes[min_subject]) else 0

    # 阈值（单科满分120分或100分）
    threshold_big = 12
    threshold_small = 6

    if max_change > threshold_big:
        reasons.append(f"{max_subject}进步明显(+{max_change:.1f}分)")
    if min_change < -threshold_big:
        reasons.append(f"{min_subject}明显下滑({min_change:.1f}分)")
    elif min_change < -threshold_small:
        reasons.append(f"{min_subject}有所退步({min_change:.1f}分)")

    # 整体评价
    progress_count = sum(1 for v in subject_changes.values() if pd.notna(v) and v > 0)
    decline_count = sum(1 for v in subject_changes.values() if pd.notna(v) and v < 0)

    if decline_count >= 4:
        reasons.append("多科目下滑，建议加强基础复习")
    elif progress_count >= 4:
        reasons.append("各科全面进步，学习状态良好")
    elif decline_count > progress_count:
        reasons.append("部分科目波动，需重点关注")
    else:
        reasons.append("整体稳定")

    return summary, "；".join(reasons) if reasons else "成绩稳定"

# 生成学生分析数据
print("生成学生分析报告...")
student_data = []
for idx, row in merged_df.iterrows():
    summary, reason = analyze_student(row)
    data = {
        '姓名': row['姓名_本次'],
        '学号': row['学号'],
        '成绩整体变化': summary,
        '波动原因推测': reason,
    }
    # 添加各科成绩
    for s in SUBJECTS:
        data[f'{s}_上次'] = row[f'{s}_上次']
        data[f'{s}_本次'] = row[f'{s}_本次']
        data[f'{s}_变化'] = row[f'{s}_变化']
    # 添加总分
    data['总分_上次'] = row['总分_上次']
    data['总分_本次'] = row['总分_本次']
    data['总分_变化'] = row['总分_变化']
    data['名次_上次'] = row['总分名次_上次']
    data['名次_本次'] = row['总分名次_本次']
    data['名次_变化'] = row['总分名次_变化']
    student_data.append(data)

df_analysis = pd.DataFrame(student_data)

# ==================== 创建Excel ====================
wb = Workbook()
wb.remove(wb.active)

# 样式定义
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ==================== Sheet 1: 学生个人分析报告 ====================
print("创建学生个人分析报告...")
ws1 = wb.create_sheet("学生个人分析报告", 0)

ws1.column_dimensions['A'].width = 12
ws1.column_dimensions['B'].width = 55
ws1.column_dimensions['C'].width = 65

ws1.append(['姓名', '成绩整体变化', '波动原因推测'])
for col in range(1, 4):
    ws1.cell(1, col).fill = header_fill
    ws1.cell(1, col).font = header_font
    ws1.cell(1, col).alignment = header_align

# 数据行
for idx, row in df_analysis.iterrows():
    ws1.append([row['姓名'], row['成绩整体变化'], row['波动原因推测']])

# 设置行高和样式
for row_idx in range(2, len(df_analysis) + 2):
    ws1.row_dimensions[row_idx].height = 45
    for col in range(1, 4):
        cell = ws1.cell(row_idx, col)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # 根据总分变化设置背景色（阈值30分）
        if col == 2:
            total_change = df_analysis.iloc[row_idx - 2]['总分_变化']
            if pd.notna(total_change):
                if total_change > 30:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif total_change < -30:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

# ==================== Sheet 2: 各科详细成绩 ====================
print("创建各科详细成绩表...")
ws2 = wb.create_sheet("各科详细成绩")

# 构建表头
detail_headers = ['姓名']
for s in SUBJECTS:
    detail_headers.extend([f'{s}↑', f'{s}↓', f'{s}变化'])
detail_headers.extend(['总分↑', '总分↓', '总分变化', '名次↑', '名次↓', '名次变化'])

ws2.append(detail_headers)
for col in range(1, len(detail_headers) + 1):
    ws2.cell(1, col).fill = header_fill
    ws2.cell(1, col).font = header_font
    ws2.cell(1, col).alignment = header_align
    ws2.column_dimensions[ws2.cell(1, col).column_letter].width = 10

# 数据
for idx, row in df_analysis.iterrows():
    row_data = [row['姓名']]
    for s in SUBJECTS:
        row_data.extend([row[f'{s}_上次'], row[f'{s}_本次'], row[f'{s}_变化']])
    row_data.extend([
        row['总分_上次'], row['总分_本次'], row['总分_变化'],
        row['名次_上次'], row['名次_本次'], row['名次_变化']
    ])
    ws2.append(row_data)

# 样式设置 - 变化列高亮
change_cols = [4, 7, 10, 13, 16, 19, 22]  # 各科变化列和名次变化列
for row_idx in range(2, len(df_analysis) + 2):
    for col in range(2, len(detail_headers) + 1):
        cell = ws2.cell(row_idx, col)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        if col in change_cols:
            value = cell.value
            if value and isinstance(value, (int, float)) and not pd.isna(value):
                if value > 0:
                    cell.font = Font(color="00B050", bold=True)
                    if col != 22:  # 名次变化列不加+号
                        cell.value = f"+{value:.1f}" if isinstance(value, float) else f"+{value}"
                elif value < 0:
                    cell.font = Font(color="FF0000", bold=True)

# ==================== Sheet 3: 班级统计分析 ====================
print("创建班级统计分析...")
ws3 = wb.create_sheet("班级统计分析")

# 标题
ws3['A1'] = '班级成绩质量分析报告'
ws3['A1'].font = Font(bold=True, size=16, color="FFFFFF")
ws3['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ws3['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws3.merge_cells('A1:H1')
ws3.row_dimensions[1].height = 30

# 统计数据表
ws3.append([''])
ws3['A3'] = '各科平均分统计'
ws3['A3'].font = Font(bold=True, size=12)
ws3.merge_cells('A3:G3')

stat_headers = ['科目', '满分', '上次平均分', '本次平均分', '平均变化', '及格率本次(%)', '优秀率本次(%)']
ws3.append(stat_headers)

for col in range(1, 8):
    ws3.cell(4, col).fill = header_fill
    ws3.cell(4, col).font = header_font
    ws3.cell(4, col).alignment = header_align

# 计算统计
stats = []
for subject in SUBJECTS:
    full_score = 120 if subject in SUBJECTS_120 else 100

    avg_prev = merged_df[f'{subject}_上次'].mean()
    avg_curr = merged_df[f'{subject}_本次'].mean()
    change = avg_curr - avg_prev

    pass_rate = (merged_df[f'{subject}_本次'] >= PASS_LINES[subject]).sum() / len(merged_df) * 100
    excel_rate = (merged_df[f'{subject}_本次'] >= EXCEL_LINES[subject]).sum() / len(merged_df) * 100

    stats.append([subject, full_score, round(avg_prev, 2), round(avg_curr, 2),
                  round(change, 2), round(pass_rate, 2), round(excel_rate, 2)])

# 总分
avg_prev_total = merged_df['总分_上次'].mean()
avg_curr_total = merged_df['总分_本次'].mean()
change_total = avg_curr_total - avg_prev_total
stats.append(['总分', TOTAL_FULL_SCORE, round(avg_prev_total, 2), round(avg_curr_total, 2),
              round(change_total, 2), '-', '-'])

for stat in stats:
    ws3.append(stat)

# 样式
for row in range(5, 11):
    for col in range(1, 8):
        cell = ws3.cell(row, col)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col == 5:  # 变化列
            if cell.value and isinstance(cell.value, (int, float)) and cell.value != 0:
                cell.font = Font(color="00B050" if cell.value > 0 else "FF0000", bold=True)

# 创建图表 1: 平均分对比
chart1 = BarChart()
chart1.title = "各科平均分对比（上次 vs 本次）"
chart1.y_axis.title = "平均分"
chart1.x_axis.title = "科目"
chart1.style = 10
chart1.height = 12
chart1.width = 20

data = Reference(ws3, min_col=3, min_row=4, max_row=9, max_col=4)
cats = Reference(ws3, min_col=1, min_row=5, max_row=9)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)

ws3.add_chart(chart1, "J3")

# 进步/退步统计
ws3['A13'] = '学生进步情况统计'
ws3['A13'].font = Font(bold=True, size=12)
ws3.merge_cells('A13:F13')

ws3.append([''])
progress_headers = ['类别', '人数', '占比(%)', '平均进步幅度']
ws3.append(progress_headers)

for col in range(1, 5):
    ws3.cell(15, col).fill = header_fill
    ws3.cell(15, col).font = header_font
    ws3.cell(15, col).alignment = header_align

# 分类统计（总分满分580分）
total_students = len(merged_df)
big_progress = (merged_df['总分_变化'] > 50).sum()  # 大幅进步（>50分）
progress = ((merged_df['总分_变化'] > 0) & (merged_df['总分_变化'] <= 50)).sum()
stable = (merged_df['总分_变化'] == 0).sum()
decline = (merged_df['总分_变化'] < 0).sum()

progress_stats = [
    ['大幅进步(>50分)', big_progress, round(big_progress/total_students*100, 2),
     round(merged_df[merged_df['总分_变化'] > 50]['总分_变化'].mean(), 2) if big_progress > 0 else 0],
    ['稳步进步(0-50分)', progress, round(progress/total_students*100, 2),
     round(merged_df[(merged_df['总分_变化'] > 0) & (merged_df['总分_变化'] <= 50)]['总分_变化'].mean(), 2) if progress > 0 else 0],
    ['退步(<0分)', decline, round(decline/total_students*100, 2),
     round(merged_df[merged_df['总分_变化'] < 0]['总分_变化'].mean(), 2) if decline > 0 else 0]
]

for stat in progress_stats:
    ws3.append(stat)

# 饼图：学生进步分布
chart2 = PieChart()
chart2.title = "学生进步情况分布"
chart2.style = 10
chart2.height = 12
chart2.width = 15

data = Reference(ws3, min_col=2, min_row=15, max_row=18)
labels = Reference(ws3, min_col=1, min_row=16, max_row=18)
chart2.add_data(data, titles_from_data=True)
chart2.set_categories(labels)

ws3.add_chart(chart2, "J21")

# ==================== Sheet 4: 进步榜和退步榜 ====================
print("创建进步榜...")
ws4 = wb.create_sheet("进步榜_退步榜")

# 进步榜
ws4['A1'] = '进步榜 TOP 20'
ws4['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws4['A1'].fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
ws4.merge_cells('A1:F1')

top_headers = ['排名', '姓名', '上次总分', '本次总分', '进步分数', '名次变化']
ws4.append(top_headers)

for col in range(1, 7):
    ws4.cell(2, col).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ws4.cell(2, col).font = Font(bold=True)
    ws4.cell(2, col).alignment = header_align

top20 = df_analysis.nlargest(20, '总分_变化')
for i, (idx, row) in enumerate(top20.iterrows(), 1):
    ws4.append([i, row['姓名'], row['总分_上次'], row['总分_本次'],
                row['总分_变化'], row['名次_变化']])

    # 前三名特殊标记
    if i <= 3:
        for col in range(1, 7):
            ws4.cell(i + 2, col).fill = PatternFill(
                start_color="FFD700" if i == 1 else "C0C0C0" if i == 2 else "CD7F32",
                end_color="FFD700" if i == 1 else "C0C0C0" if i == 2 else "CD7F32",
                fill_type="solid"
            )

# 需要关注学生（退步较大）
ws4['A25'] = '需要关注学生（退步较大）'
ws4['A25'].font = Font(bold=True, size=14, color="FFFFFF")
ws4['A25'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
ws4.merge_cells('A25:F25')

ws4.append(top_headers)
for col in range(1, 7):
    ws4.cell(26, col).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws4.cell(26, col).font = Font(bold=True)
    ws4.cell(26, col).alignment = header_align

bottom_students = df_analysis[df_analysis['总分_变化'] < 0].nsmallest(10, '总分_变化')
if len(bottom_students) > 0:
    for i, (idx, row) in enumerate(bottom_students.iterrows(), 1):
        ws4.append([i, row['姓名'], row['总分_上次'], row['总分_本次'],
                    row['总分_变化'], row['名次_变化']])

# ==================== 保存文件 ====================
print("\n保存Excel文件...")
output_file = '成绩分析报告_2025.xlsx'
wb.save(output_file)

print(f"\n[OK] Excel文件创建成功！")
print(f"文件名：{output_file}")
print(f"\n包含以下Sheet：")
print(f"  1. 学生个人分析报告 - {len(df_analysis)}名学生个性化分析")
print(f"  2. 各科详细成绩 - 所有科目详细对比")
print(f"  3. 班级统计分析 - 包含2个图表")
print(f"  4. 进步榜_退步榜 - TOP20及需要关注学生")
print(f"\n关键发现：")
print(f"  - 班级平均总分：{avg_prev_total:.2f} -> {avg_curr_total:.2f} (变化{change_total:+.2f}分)")
print(f"  - 大幅进步学生(>50分)：{big_progress}人 ({big_progress/total_students*100:.1f}%)")
print(f"  - 退步学生：{decline}人 ({decline/total_students*100:.1f}%)")
print(f"\n注：各科按原始满分制统计（语数英科120分，社会100分）")
